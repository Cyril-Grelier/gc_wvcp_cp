"""
Generate a xlsx file containing results of given methods from json files.

Follow the main function to manage the files and parameters
"""
import re
import statistics
from glob import glob
from typing import Any, Iterable
import json
import os

import pandas as pd  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Alignment, Font  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore


# When a best score is equal to the best known score : red
COLOR_BEST = "ff0000"
# When a best score is proven optimal : blue
COLOR_OPTIMAL = "0000ff"
# When a best score is a new proven optimal : green
COLOR_NEW_OPTIMAL = "00ff00"
# When a best score is better than the best known score : orange
COLOR_NEW_BEST = "FF7F00"


COLOR_GAP1 = "488f31"
COLOR_GAP2 = "de425b"


def main():
    """
    Choose the methods and instances and create the xlsx file
    """
    # Add method name and repertory of data of each method
    methods: list[tuple(str, str)] = [
        # E0 original vs reduced
        # ("primal original", "primal_h1_none_no_bounds_original"),
        # ("primal reduced", "primal_h1_none_no_bounds"),
        # ("dual original", "dual_h1_none_no_bounds_original"),
        # ("dual reduced", "dual_h1_none_no_bounds"),
        # ("joint original", "joint_h1_none_no_bounds_original"),
        # ("joint reduced", "joint_h1_none_no_bounds"),
        # E1 color/score bounds
        # ("primal", "primal_h1_none_no_bounds"),
        # ("primal lb color", "primal_h1_none_lb_color"),
        # ("primal ub color", "primal_h1_none_ub_color"),
        # ("primal lb score", "primal_h1_none_lb_score"),
        # ("primal ub score", "primal_h1_none_ub_score"),
        # ("primal bounds", "primal_h1_none_all_bounds"),
        # ("dual", "dual_h1_none_no_bounds"),
        # ("dual lb color", "dual_h1_none_lb_color"),
        # ("dual ub color", "dual_h1_none_ub_color"),
        # ("dual lb score", "dual_h1_none_lb_score"),
        # ("dual ub score", "dual_h1_none_ub_score"),
        # ("dual bounds", "dual_h1_none_all_bounds"),
        # ("joint", "joint_h1_none_no_bounds"),
        # ("joint lb color", "joint_h1_none_lb_color"),
        # ("joint ub color", "joint_h1_none_ub_color"),
        # ("joint lb score", "joint_h1_none_lb_score"),
        # ("joint ub score", "joint_h1_none_ub_score"),
        # ("joint bounds", "joint_h1_none_all_bounds"),
        # # # E2 none/static/dynamic
        # ("primal", "primal_h1_none_no_bounds"),
        # ("primal S/D", "primal_h1_DG_MLS_no_bounds"),
        # ("dual", "dual_h1_none_no_bounds"),
        # # ("joint", "joint_h1_none_no_bounds"),
        # ("joint S/D", "joint_h1_DG_LS_MLS_no_bounds"),
        # # E3 none/static/dynamic + bounds
        # ("primal", "primal_h1_none_no_bounds"),
        # ("primal bounds", "primal_h1_none_all_bounds"),
        # ("primal S/D", "primal_h1_DG_MLS_no_bounds"),
        # ("primal S/D bounds", "primal_h1_DG_MLS_all_bounds"),
        # ("dual", "dual_h1_none_no_bounds"),
        # ("dual bounds", "dual_h1_none_all_bounds"),
        # ("joint", "joint_h1_none_no_bounds"),
        # ("joint bounds", "joint_h1_none_all_bounds"),
        # ("joint S/D", "joint_h1_DG_LS_MLS_no_bounds"),
        # ("joint S/D bounds", "joint_h1_DG_LS_MLS_all_bounds"),
        # # 1h run, 8Go RAM, 10 CPU
        # # E4 parallelism
        ("primal S/D bounds parallel", "primal_h1_DG_MLS_all_bounds_parallel"),
        ("dual bounds parallel", "dual_h1_none_all_bounds_parallel"),
        ("joint S/D bounds parallel", "joint_h1_DG_LS_MLS_all_bounds_parallel"),
    ]

    for i, method in enumerate(methods):
        methods[i] = (method[0], "cp_1h_all/" + method[1])

    problem = "gcp"
    problem = "wvcp"

    # Choose the set of instances
    instances_set = ("pxx", "pxx")
    instances_set = ("rxx", "rxx")
    instances_set = ("DIMACS_non_optimal", "dimacs_no")
    instances_set = ("DIMACS_optimal", "dimacs_o")
    instances_set = ("instance_list_gcp", "all")
    instances_set = ("instance_list_wvcp", "all")
    # instances_set = ("../instances_try_hard", "to_opti")

    instance_type = "reduced"
    instance_type = "original"

    # output_file = f"xlsx_files/E0_original_vs_reduced_{instances_set[1]}.xlsx"
    # output_file = f"xlsx_files/E1_bounds_{instances_set[1]}.xlsx"
    # output_file = f"xlsx_files/E2_static_dynamic_{instances_set[1]}.xlsx"
    # output_file = f"xlsx_files/E3_SD_bounds_{instances_set[1]}.xlsx"
    output_file = f"xlsx_files/E4_parallel_{instances_set[1]}.xlsx"

    with open(f"instances/{instances_set[0]}.txt", "r", encoding="utf8") as file:
        instances = [i[:-1] for i in file.readlines()]

    table = Table(
        methods=methods,
        instances=instances,
        problem=problem,
        instance_type=instance_type,
    )
    table.to_xlsx(output_file)
    print(output_file)


class Method:
    def __init__(
        self, name: str, repertory: str, instance_name, instance_type: str
    ) -> None:
        self.name: str = name
        self.repertory: str = repertory
        self.flat_time: float = float("inf")
        self.solve_time: float = float("inf")
        self.score: float = float("inf")
        self.optimality_time: float = float("inf")
        self.optimal: bool = False
        self.objective_bound: float = float("inf")
        self.failures: float = float("inf")

        # load data
        json_file = f"outputs/{repertory}/{instance_name}.json"
        if os.path.exists(json_file):
            with open(json_file, "r", encoding="utf8") as file:
                for line in file.readlines():
                    l_json = json.loads(line)
                    if l_json["type"] == "statistics":
                        if "flatTime" in l_json["statistics"]:
                            self.flat_time = int(l_json["statistics"]["flatTime"])
                            if self.flat_time > 3600:
                                print(instance_name, "flat time > 3600")
                        else:
                            if self.optimal:
                                self.optimality_time = int(
                                    l_json["statistics"]["solveTime"]
                                )
                            else:
                                self.solve_time = int(l_json["statistics"]["solveTime"])
                            try:
                                self.objective_bound = l_json["statistics"][
                                    "objectiveBound"
                                ]
                            except:
                                pass
                            try:
                                self.failures = l_json["statistics"]["failures"]
                            except:
                                self.objective_bound = int(self.objective_bound)
                    elif l_json["type"] == "solution":
                        if "x_o" in l_json["output"]["json"]:
                            self.score = l_json["output"]["json"]["x_o"]
                        else:
                            self.score = l_json["output"]["json"]["yx_o"]
                    elif l_json["type"] == "status":
                        self.optimal = l_json["status"] == "OPTIMAL_SOLUTION"


class Gap:
    def __init__(self, m_1: Method, m_2: Method) -> None:
        inf = float("inf")
        self.score_diff: float = inf
        self.time_diff: float = inf
        self.time_optim_diff = inf
        self.optim_diff = inf
        self.better_method_score: str = ""
        self.better_method_time: str = ""
        self.better_method_optim: str = ""
        self.better_method_optim_time: str = ""

        if m_1.optimality_time == inf and m_2.optimality_time != inf:
            self.better_method_optim = m_2.name
            self.optim_diff = -1
        if m_2.optimality_time == inf and m_1.optimality_time != inf:
            self.better_method_optim = m_1.name
            self.optim_diff = 1

        if m_1.score == inf and m_2.score == inf:
            return
        if m_1.score == inf:
            self.better_method_score = m_2.name
            return
        if m_2.score == inf:
            self.better_method_score = m_1.name
            return

        self.score_diff = round(m_1.score - m_2.score, 1)
        if self.score_diff < 0:
            self.better_method_score = m_1.name
            return
        if self.score_diff > 0:
            self.better_method_score = m_2.name
            return

        # same score - diff time
        self.time_diff = round(m_1.solve_time - m_2.solve_time, 1)
        if self.time_diff < -1:
            self.better_method_time = m_1.name
        if self.time_diff > 1:
            self.better_method_time = m_2.name

        if m_1.optimality_time == inf and m_2.optimality_time == inf:
            return

        self.optim_diff = 0
        self.time_optim_diff = round(m_1.optimality_time - m_2.optimality_time, 1)
        if self.time_optim_diff < -1:
            self.better_method_optim_time = m_1.name
        if self.time_optim_diff > 1:
            self.better_method_optim_time = m_2.name


class Instance:
    """Store the results of all given method on one instance"""

    def __init__(
        self,
        name: str,
        methods: list[tuple[str, str]],
        problem: str,
        instance_type: str,
        gaps: list[tuple[str, str]],
    ) -> None:
        self.name: str = name
        self.nb_vertices: int
        self.nb_edges: int
        self.best_known_score: int
        self.optimal: bool

        self.methods: dict[str, Method] = {
            m_name: Method(m_name, repertory, name, instance_type)
            for m_name, repertory in methods
        }
        self.methods_names: list[str] = [m_name for m_name, _ in methods]
        self.gaps: dict[tuple[str, str], Gap] = dict()

        # get informations on the instance
        self.nb_vertices, self.nb_edges, self.density = get_density(
            name, instance_type, problem
        )
        self.best_known_score, self.optimal = get_best_known_score(name, problem)

        # compute gap between methods
        for m_1, m_2 in gaps:
            self.gaps[(m_1, m_2)] = Gap(self.methods[m_1], self.methods[m_2])


class Table:
    """Representation of the data table"""

    def __init__(
        self,
        methods: list[tuple[str, str]],
        instances: list[str],
        problem: str,
        instance_type: str,
    ) -> None:

        self.methods_names: list[str] = [m_name for m_name, _ in methods]

        self.gaps: list[tuple[str, str]] = []
        # pylint: disable=consider-using-enumerate
        for i in range(len(self.methods_names)):
            for j in range(i + 1, len(self.methods_names)):
                self.gaps.append((self.methods_names[i], self.methods_names[j]))

        self.instances: list[Instance] = [
            Instance(instance, methods, problem, instance_type, self.gaps)
            for instance in instances
        ]

        self.nb_flat = {
            m: sum(
                1
                for instance in self.instances
                if instance.methods[m].flat_time != float("inf")
            )
            for m in self.methods_names
        }

        self.nb_solved = {
            m: sum(
                1
                for instance in self.instances
                if instance.methods[m].solve_time != float("inf")
            )
            for m in self.methods_names
        }

        self.nb_optim = {
            m: sum(1 for instance in self.instances if instance.methods[m].optimal)
            for m in self.methods_names
        }

        self.nb_best_score = {
            m: sum(
                1
                for instance in self.instances
                if instance.methods[m].score == instance.best_known_score
            )
            for m in self.methods_names
        }

        self.nb_gaps_m1_best_score: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_score == m1)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m2_best_score: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_score == m2)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m1_best_time: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_time == m1)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m2_best_time: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_time == m2)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m1_best_optim: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_optim == m1)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m2_best_optim: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_optim == m2)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m1_best_optim_time: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_optim_time == m1)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m2_best_optim_time: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_optim_time == m2)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }

        self.ranks: dict[str, list[list[str]]] = dict()
        self.nb_ranks: dict[str, list[int]] = {
            m: [0] * (len(self.methods_names) + 1) for m in self.methods_names
        }
        for instance in self.instances:
            methods_ = list(self.methods_names)
            keys = {
                m: (
                    instance.methods[m].score,
                    not instance.methods[m].optimal,
                    instance.methods[m].optimality_time,
                    instance.methods[m].solve_time,
                )
                for m in methods_
            }
            methods_.sort(key=lambda m: keys[m])
            ranks: list[list[str]] = [[] for _ in range(len(methods_) + 1)]
            last_rank = 0
            for i, m in enumerate(methods_):
                if keys[m] == (float("inf"), True, float("inf"), float("inf")):
                    last_rank = len(ranks) - 1
                    ranks[last_rank].append(m)
                elif not ranks[last_rank]:
                    ranks[last_rank].append(m)
                elif keys[ranks[last_rank][0]] == keys[m]:
                    ranks[last_rank].append(m)
                else:
                    last_rank = i
                    ranks[last_rank].append(m)
                self.nb_ranks[m][last_rank] += 1
            self.ranks[instance.name] = ranks

    def __repr__(self) -> str:
        return "\n".join([str(instance) for instance in self.instances])

    def table_results(self, workbook: Workbook):
        sheet = workbook.active
        sheet.title = "results"
        # first row
        # first columns are the instances informations then the methods names
        instance_info = [
            "instance",
            "|V|",
            "|E|",
            "density",
            "BKS",
            # "optim",
        ]
        columns_info = [
            "score",
            "flat(s)",
            "solve(s)",
            "time(s)",
            "LB",
            "failures",
        ]
        line: list[int | str | float] = ["" for _ in instance_info]
        line += [m for m in self.methods_names for _ in columns_info]
        sheet.append(line)
        # merge first row for methods names
        for i in range(len(self.methods_names)):
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # second row
        # instance informations then columns info
        line = list(instance_info)
        for _ in self.methods_names:
            line += list(columns_info)
        sheet.append(line)

        # # merge 2 firsts lines for instances informations
        # for i in range(len(instance_info)):
        #     sheet.merge_cells(
        #         start_row=1, end_row=2, start_column=i + 1, end_column=i + 1
        #     )

        # body of the table
        # first columns are the instance info then the scores, times,... for each methods
        for instance in self.instances:
            line = [
                instance.name,
                instance.nb_vertices,
                instance.nb_edges,
                instance.density,
                str(instance.best_known_score) + "*"
                if instance.optimal
                else instance.best_known_score,
                # instance.optimal,
            ]
            for m in self.methods_names:
                method = instance.methods[m]
                line += [
                    method.score,
                    method.flat_time,
                    method.solve_time,
                    method.optimality_time
                    if method.optimality_time != float("inf")
                    else "tl",
                    method.objective_bound,
                    method.failures,
                ]
            sheet.append(line)
            for col, m in enumerate(self.methods_names):
                column_best_score = len(instance_info) + 1 + len(columns_info) * col
                cell_best_score = sheet.cell(sheet.max_row, column_best_score)
                if cell_best_score.value == float("inf"):
                    cell_best_score.value = "-"
                    continue
                val_best_score = int(cell_best_score.value)
                if val_best_score == instance.best_known_score:
                    cell_best_score.font = Font(bold=True, color=COLOR_BEST)
                elif val_best_score < instance.best_known_score:
                    cell_best_score.font = Font(
                        bold=True, color=COLOR_NEW_BEST, underline="single"
                    )
                if instance.methods[m].optimal and instance.optimal:
                    cell_best_score.value = str(cell_best_score.value) + "*"
                    cell_best_score.font = Font(bold=True, color=COLOR_OPTIMAL)
                elif instance.methods[m].optimal:
                    cell_best_score.value = str(cell_best_score.value) + "*"
                    cell_best_score.font = Font(
                        bold=True, color=COLOR_NEW_OPTIMAL, underline="single"
                    )

        # footer

        # # nb flat
        # line = ["nb flat done"] * len(instance_info)
        # for m in self.methods_names:
        #     line += [f"{self.nb_flat[m]}/{len(self.instances)}"] * len(columns_info)
        # sheet.append(line)
        # # merge footer
        # sheet.merge_cells(
        #     start_row=sheet.max_row,
        #     end_row=sheet.max_row,
        #     start_column=1,
        #     end_column=len(instance_info),
        # )
        # for i, _ in enumerate(self.methods_names):
        #     sheet.merge_cells(
        #         start_row=sheet.max_row,
        #         end_row=sheet.max_row,
        #         start_column=len(instance_info) + 1 + len(columns_info) * i,
        #         end_column=len(instance_info) + len(columns_info) * (i + 1),
        #     )

        # # nb solve
        # line = ["nb solved"] * len(instance_info)
        # for m in self.methods_names:
        #     line += [f"{self.nb_solved[m]}/{len(self.instances)}"] * len(columns_info)
        # sheet.append(line)
        # # merge footer
        # sheet.merge_cells(
        #     start_row=sheet.max_row,
        #     end_row=sheet.max_row,
        #     start_column=1,
        #     end_column=len(instance_info),
        # )
        # for i, _ in enumerate(self.methods_names):
        #     sheet.merge_cells(
        #         start_row=sheet.max_row,
        #         end_row=sheet.max_row,
        #         start_column=len(instance_info) + 1 + len(columns_info) * i,
        #         end_column=len(instance_info) + len(columns_info) * (i + 1),
        #     )

        # nb best known score
        line = ["nb best known score"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_best_score[m]}/{len(self.instances)}"] * len(
                columns_info
            )
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # nb optimal
        line = ["nb proved optimal"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_optim[m]}/{len(self.instances)}"] * len(columns_info)
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) + 1 > column_widths[i]:
                        column_widths[i] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = column_width

        # Freeze row and columns
        sheet.freeze_panes = sheet["F3"]

    def table_results_to_latex(self, file_name: str):

        instance_info = [
            "instance",
            # "|V|",
            # "|E|",
            # "density",
            "BKS",
        ]
        columns_info = [
            "score",
            # "flat(s)",
            # "solve(s)",
            "LB",
            "time(s)",
            # "failures",
        ]

        table = "\\begin{table}[h]\n"
        table += "\\resizebox{\\linewidth}{!}{\n"
        table += "\\centering\n"
        table += (
            "\\begin{tabular}{lr|"
            + (("r" * (len(columns_info)) + "|") * (len(self.methods_names) - 1))
            + "r" * (len(columns_info))
            + "}\n"
        )
        table += "\\toprule\n"

        table += "&" * len(instance_info)
        for m in self.methods_names[:-1]:
            table += "\multicolumn{" + str(len(columns_info)) + "}{c|}{" + m + "} &"
        table += (
            "\multicolumn{"
            + str(len(columns_info))
            + "}{c}{"
            + self.methods_names[-1]
            + "} \\\\\n"
        )
        for info in instance_info:
            table += info + "&"
        for m in self.methods_names[:-1]:
            for info in columns_info:
                table += info + "&"
        for info in columns_info[:-1]:
            table += info + "&"
        table += columns_info[-1] + "\\\\\n"
        table += "\\midrule\n"

        # body
        for instance in self.instances:
            if "_" in instance.name:
                table += instance.name.replace("_", "\\_") + "&"
            else:
                table += instance.name + "&"
            # table += str(instance.nb_vertices) + "&"
            # table += str(instance.nb_edges) + "&"
            # table += str(instance.density) + "&"
            table += str(instance.best_known_score)
            if instance.optimal:
                table += "*&"
            else:
                table += "&"

            for m in self.methods_names[:-1]:
                method = instance.methods[m]
                if method.score != float("inf"):
                    if method.optimal and not instance.optimal:
                        table += "\\textbf{\\underline{" + str(method.score) + "*}}&"
                    elif method.optimal:
                        table += "\\textbf{" + str(method.score) + "*}&"
                    elif method.score < instance.best_known_score:
                        table += "\\underline{" + str(method.score) + "}&"
                    else:
                        table += str(method.score) + "&"
                else:
                    table += "-&"
                # table += str(method.flat_time) + "&"
                # table += str(method.solve_time) + "&"
                if method.objective_bound != float("inf"):
                    table += str(method.objective_bound) + "&"
                else:
                    table += " &"
                if method.optimality_time != float("inf"):
                    table += str(method.optimality_time) + "&"
                else:
                    table += "tl&"
                # table += str(method.failures) + "&"

            method = instance.methods[self.methods_names[-1]]
            if method.score != float("inf"):
                if method.optimal and not instance.optimal:
                    table += "\\textbf{\\underline{" + str(method.score) + "*}}&"
                elif method.optimal:
                    table += "\\textbf{" + str(method.score) + "*}&"
                elif method.score < instance.best_known_score:
                    table += "\\underline{" + str(method.score) + "}&"
                else:
                    table += str(method.score) + "&"
            else:
                table += "-&"
            # table += str(method.flat_time) + "&"
            # table += str(method.solve_time) + "&"
            if method.objective_bound != float("inf"):
                table += str(method.objective_bound) + "&"
            else:
                table += " &"
            if method.optimality_time != float("inf"):
                table += str(method.optimality_time) + "\\\\\n"
            else:
                table += "tl\\\\\n"
            # table += str(method.objective_bound) + "&"
            # table += str(method.failures) + "&"

        table += "\\bottomrule\n"

        # footer

        # nb best known score
        table += "\multicolumn{2}{c|}{nb bks}&"
        for m in self.methods_names[:-1]:
            table += (
                "\multicolumn{"
                + str(len(columns_info))
                + "}{c|}{"
                + str(self.nb_best_score[m])
                + "/"
                + str(len(self.instances))
                + "} &"
            )
        table += (
            "\multicolumn{"
            + str(len(columns_info))
            + "}{c}{"
            + str(self.nb_best_score[self.methods_names[-1]])
            + "/"
            + str(len(self.instances))
            + "} \\\\\n"
        )

        # nb optimal
        table += "\multicolumn{2}{c|}{nb optim}&"
        for m in self.methods_names[:-1]:
            table += (
                "\multicolumn{"
                + str(len(columns_info))
                + "}{c|}{"
                + str(self.nb_optim[m])
                + "/"
                + str(len(self.instances))
                + "} &"
            )
        table += (
            "\multicolumn{"
            + str(len(columns_info))
            + "}{c}{"
            + str(self.nb_optim[self.methods_names[-1]])
            + "/"
            + str(len(self.instances))
            + "} \\\\\n"
        )

        # end
        table += "\\bottomrule\n"
        table += "\\end{tabular}\n"
        table += "}\n"
        table += "\\end{table}\n"

        with open(file_name, "w", encoding="utf8") as file:
            file.write(table)

    def table_gaps(self, workbook: Workbook):
        sheet = workbook.create_sheet("gaps")

        instance_info = ["instance"]
        columns_info = ["gap score", "gap time", "gap optim", "gap optim time"]
        # first row
        line: list[int | str | float] = list(instance_info)
        line += [
            f"m1 : {m1} \n- m2 : {m2}" for m1, m2 in self.gaps for _ in columns_info
        ]
        sheet.append(line)
        # merge first row for methods names
        for i in range(len(self.gaps)):
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # second row
        # instance informations then columns info
        line = list(instance_info)
        for i in range(len(self.gaps)):
            line += list(columns_info)
        sheet.append(line)

        # merge 2 firsts lines for instances informations
        for i in range(len(instance_info)):
            sheet.merge_cells(
                start_row=1, end_row=2, start_column=i + 1, end_column=i + 1
            )

        # body of the table
        # first columns are the instance info then the gaps for each methods
        for instance in self.instances:
            line = [instance.name]
            for m1, m2 in self.gaps:
                gap = instance.gaps[(m1, m2)]
                line += [
                    gap.score_diff,
                    gap.time_diff,
                    gap.optim_diff,
                    gap.time_optim_diff,
                ]
            sheet.append(line)

            # add color
            for col in range(len(self.gaps)):
                column_diff_score = len(instance_info) + 1 + len(columns_info) * col
                column_diff_time = column_diff_score + 1
                column_diff_optim = column_diff_time + 1
                column_diff_optim_time = column_diff_time + 1
                cell_diff_score = sheet.cell(sheet.max_row, column_diff_score)
                cell_diff_time = sheet.cell(sheet.max_row, column_diff_time)
                cell_diff_optim = sheet.cell(sheet.max_row, column_diff_optim)
                cell_diff_optim_time = sheet.cell(sheet.max_row, column_diff_optim_time)
                if cell_diff_score.value == float("inf"):
                    continue
                val_diff_score = float(cell_diff_score.value)
                val_diff_time = float(cell_diff_time.value)
                val_diff_optim = float(cell_diff_optim.value)
                val_diff_optim_time = float(cell_diff_optim_time.value)

                color_score = "808080"  # gray
                color_time = "808080"  # gray
                color_optim = "808080"  # gray
                color_optim_time = "808080"  # gray
                if val_diff_score > 0:
                    # 1 better
                    color_score = COLOR_GAP1
                elif val_diff_score < 0:
                    # 2 better
                    color_score = COLOR_GAP2
                else:
                    if val_diff_time > 1:
                        color_time = COLOR_GAP1
                    elif val_diff_time < -1:
                        color_time = COLOR_GAP2
                    if val_diff_optim == 1:
                        color_optim = COLOR_GAP1
                    elif val_diff_optim == -1:
                        color_optim = COLOR_GAP2
                    if val_diff_optim_time > 1:
                        color_optim_time = COLOR_GAP1
                    elif val_diff_optim_time < -1:
                        color_optim_time = COLOR_GAP2
                cell_diff_score.font = Font(bold=True, color=color_score)
                cell_diff_time.font = Font(bold=True, color=color_time)
                cell_diff_optim.font = Font(bold=True, color=color_optim)
                cell_diff_optim_time.font = Font(bold=True, color=color_optim_time)

        # footer
        line = ["m1 better"] * len(instance_info)
        for m1, m2 in self.gaps:
            line += [
                self.nb_gaps_m1_best_score[(m1, m2)],
                self.nb_gaps_m1_best_time[(m1, m2)],
                self.nb_gaps_m1_best_optim[(m1, m2)],
                self.nb_gaps_m1_best_optim_time[(m1, m2)],
            ]
        sheet.append(line)
        for col in range(1, sheet.max_column + 1):
            sheet.cell(sheet.max_row, col).font = Font(bold=True, color=COLOR_GAP2)

        line = ["m2 better"] * len(instance_info)
        for m1, m2 in self.gaps:
            line += [
                self.nb_gaps_m2_best_score[(m1, m2)],
                self.nb_gaps_m2_best_time[(m1, m2)],
                self.nb_gaps_m2_best_optim[(m1, m2)],
                self.nb_gaps_m2_best_optim_time[(m1, m2)],
            ]
        sheet.append(line)

        for col in range(1, sheet.max_column + 1):
            sheet.cell(sheet.max_row, col).font = Font(bold=True, color=COLOR_GAP1)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) + 1 > column_widths[i]:
                        column_widths[i] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = column_width

        # Freeze row and columns
        sheet.freeze_panes = sheet["B3"]

    def table_comparaison_score(self, workbook: Workbook):
        sheet = workbook.create_sheet("score")
        # first row with the methods names
        line: list[str | int] = [""]
        for m1 in self.methods_names:
            line.append(m1)
        sheet.append(line)

        # add data, first method name then nb significative gaps
        for row, m1 in enumerate(self.methods_names):
            m1 = self.methods_names[row]
            line = []
            line.append(m1)
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    line.append("-")
                else:
                    if row < col:
                        line.append(self.nb_gaps_m1_best_score[(m1, m2)])
                    else:
                        line.append(self.nb_gaps_m2_best_score[(m2, m1)])
            sheet.append(line)

            # bold best scores
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    continue
                cell_m1 = sheet.cell(row + 2, col + 2)
                if (
                    row < col
                    and (
                        self.nb_gaps_m1_best_score[(m1, m2)]
                        > self.nb_gaps_m2_best_score[(m1, m2)]
                    )
                ) or (
                    row > col
                    and (
                        self.nb_gaps_m2_best_score[(m2, m1)]
                        > self.nb_gaps_m1_best_score[(m2, m1)]
                    )
                ):
                    cell_m1.font = Font(bold=True)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rotate the first row
        for col in range(len(self.methods_names)):
            met_name_cell = sheet.cell(1, col + 2)
            met_name_cell.alignment = Alignment(textRotation=90)

        # add info
        line = [""] * (len(self.methods_names) + 1)
        line.append("number of times the method on the row is find a better solution")
        sheet.append(line)

        # Set optimal width
        column_widths: list[int] = []
        row_: Iterable[Iterable[Any]]
        for row_ in sheet:
            for col, cell in enumerate(row_):
                if len(column_widths) > col:
                    if len(str(cell.value)) + 1 > column_widths[col]:
                        column_widths[col] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_comparaison_time(self, workbook: Workbook):
        sheet = workbook.create_sheet("time")
        # first row with the methods names
        line: list[str | int] = [""]
        for m1 in self.methods_names:
            line.append(m1)
        sheet.append(line)

        # add data, first method name then nb significative gaps
        for row, m1 in enumerate(self.methods_names):
            m1 = self.methods_names[row]
            line = []
            line.append(m1)
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    line.append("-")
                else:
                    if row < col:
                        line.append(self.nb_gaps_m1_best_time[(m1, m2)])
                    else:
                        line.append(self.nb_gaps_m2_best_time[(m2, m1)])
            sheet.append(line)

            # bold best scores
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    continue
                cell_m1 = sheet.cell(row + 2, col + 2)
                if (
                    row < col
                    and (
                        self.nb_gaps_m1_best_time[(m1, m2)]
                        > self.nb_gaps_m2_best_time[(m1, m2)]
                    )
                ) or (
                    row > col
                    and (
                        self.nb_gaps_m2_best_time[(m2, m1)]
                        > self.nb_gaps_m1_best_time[(m2, m1)]
                    )
                ):
                    cell_m1.font = Font(bold=True)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rotate the first row
        for col in range(len(self.methods_names)):
            met_name_cell = sheet.cell(1, col + 2)
            met_name_cell.alignment = Alignment(textRotation=90)

        # add info
        line = [""] * (len(self.methods_names) + 1)
        line.append(
            "number of times the method on the row is faster to find best solution (if they found the same score)"
        )
        sheet.append(line)

        # Set optimal width
        column_widths: list[int] = []
        row_: Iterable[Iterable[Any]]
        for row_ in sheet:
            for col, cell in enumerate(row_):
                if len(column_widths) > col:
                    if len(str(cell.value)) + 1 > column_widths[col]:
                        column_widths[col] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_comparaison_optim(self, workbook: Workbook):
        sheet = workbook.create_sheet("optim")
        # first row with the methods names
        line: list[str | int] = [""]
        for m1 in self.methods_names:
            line.append(m1)
        sheet.append(line)

        # add data, first method name then nb significative gaps
        for row, m1 in enumerate(self.methods_names):
            m1 = self.methods_names[row]
            line = []
            line.append(m1)
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    line.append("-")
                else:
                    if row < col:
                        line.append(self.nb_gaps_m1_best_optim[(m1, m2)])
                    else:
                        line.append(self.nb_gaps_m2_best_optim[(m2, m1)])
            sheet.append(line)

            # bold best scores
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    continue
                cell_m1 = sheet.cell(row + 2, col + 2)
                if (
                    row < col
                    and (
                        self.nb_gaps_m1_best_optim[(m1, m2)]
                        > self.nb_gaps_m2_best_optim[(m1, m2)]
                    )
                ) or (
                    row > col
                    and (
                        self.nb_gaps_m2_best_optim[(m2, m1)]
                        > self.nb_gaps_m1_best_optim[(m2, m1)]
                    )
                ):
                    cell_m1.font = Font(bold=True)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rotate the first row
        for col in range(len(self.methods_names)):
            met_name_cell = sheet.cell(1, col + 2)
            met_name_cell.alignment = Alignment(textRotation=90)

        # add info
        line = [""] * (len(self.methods_names) + 1)
        line.append(
            "number of times the method on the row find optimality while the one on the column not"
        )
        sheet.append(line)

        # Set optimal width
        column_widths: list[int] = []
        row_: Iterable[Iterable[Any]]
        for row_ in sheet:
            for col, cell in enumerate(row_):
                if len(column_widths) > col:
                    if len(str(cell.value)) + 1 > column_widths[col]:
                        column_widths[col] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_comparaison_optim_time(self, workbook: Workbook):
        sheet = workbook.create_sheet("optim time")
        # first row with the methods names
        line: list[str | int] = [""]
        for m1 in self.methods_names:
            line.append(m1)
        sheet.append(line)

        # add data, first method name then nb significative gaps
        for row, m1 in enumerate(self.methods_names):
            m1 = self.methods_names[row]
            line = []
            line.append(m1)
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    line.append("-")
                else:
                    if row < col:
                        line.append(self.nb_gaps_m1_best_optim_time[(m1, m2)])
                    else:
                        line.append(self.nb_gaps_m2_best_optim_time[(m2, m1)])
            sheet.append(line)

            # bold best scores
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    continue
                cell_m1 = sheet.cell(row + 2, col + 2)
                if (
                    row < col
                    and (
                        self.nb_gaps_m1_best_optim_time[(m1, m2)]
                        > self.nb_gaps_m2_best_optim_time[(m1, m2)]
                    )
                ) or (
                    row > col
                    and (
                        self.nb_gaps_m2_best_optim_time[(m2, m1)]
                        > self.nb_gaps_m1_best_optim_time[(m2, m1)]
                    )
                ):
                    cell_m1.font = Font(bold=True)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rotate the first row
        for col in range(len(self.methods_names)):
            met_name_cell = sheet.cell(1, col + 2)
            met_name_cell.alignment = Alignment(textRotation=90)

        # add info
        line = [""] * (len(self.methods_names) + 1)
        line.append(
            "number of times the method on the row is faster to prove optimality (if they found the same score)"
        )
        sheet.append(line)

        # Set optimal width
        column_widths: list[int] = []
        row_: Iterable[Iterable[Any]]
        for row_ in sheet:
            for col, cell in enumerate(row_):
                if len(column_widths) > col:
                    if len(str(cell.value)) + 1 > column_widths[col]:
                        column_widths[col] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_ranks(self, workbook: Workbook):
        sheet = workbook.create_sheet("ranks")
        # first row
        line = [""]
        for i in range(len(self.methods_names)):
            line.append(f"rank {i+1}")
        line.append("rank unsolved")
        sheet.append(line)

        # data
        for instance in self.instances:
            line = [instance.name]
            for i, rank in enumerate(self.ranks[instance.name]):
                txt = ""
                for m in rank:
                    txt += f"{m}\n"
                if rank and (i != (len(self.ranks[instance.name]) - 1)):
                    m = rank[0]
                    txt += (
                        f"s={instance.methods[m].score} "
                        f"o={instance.methods[m].optimal} "
                        f"ot={instance.methods[m].optimality_time} "
                        f"st={instance.methods[m].solve_time}"
                    )
                line.append(txt)
            sheet.append(line)

        # Set alignment
        for col in sheet.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for row, cell in enumerate(row):
                if len(column_widths) > row:
                    if len(str(cell.value)) + 1 > column_widths[row]:
                        column_widths[row] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_summary_ranks(self, workbook: Workbook):
        sheet = workbook.create_sheet("summary_ranks")
        # first row
        line: list[int | str | float] = [""]
        for i in range(len(self.methods_names)):
            line.append(f"rank {i+1}")
        line.append("rank unsolved")
        sheet.append(line)

        methods = list(self.methods_names)
        scores = {m: 0 for m in methods}
        for m in methods:
            for i, nb in enumerate(self.nb_ranks[m]):
                scores[m] += nb * (len(self.nb_ranks[m]) - i)
        # methods.sort(key=lambda m: scores[m], reverse=True)
        # data
        for m in methods:
            line = [m]
            for nb_rank in self.nb_ranks[m]:
                line.append(nb_rank)
            sheet.append(line)

        # Set alignment
        for col in sheet.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for row, cell in enumerate(row):
                if len(column_widths) > row:
                    if len(str(cell.value)) + 1 > column_widths[row]:
                        column_widths[row] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    # def table_instances(self, workbook: Workbook):
    #     sheet = workbook.create_sheet("instances")
    #     # first row
    #     line = [""]
    #     for i in range(len(self.methods_names)):
    #         line.append(f"rank {i+1}")
    #     line.append("rank unsolved")
    #     sheet.append(line)

    #     # data
    #     for instance in self.instances:
    #         line = [instance.name]
    #         for i, rank in enumerate(self.ranks[instance.name]):
    #             txt = ""
    #             for m in rank:
    #                 txt += f"{m}\n"
    #             if rank and (i != (len(self.ranks[instance.name]) - 1)):
    #                 m = rank[0]
    #                 txt += (
    #                     f"s={instance.methods[m].score} "
    #                     f"o={instance.methods[m].optimal} "
    #                     f"ot={instance.methods[m].optimality_time} "
    #                     f"st={instance.methods[m].solve_time}"
    #                 )
    #             line.append(txt)
    #         sheet.append(line)

    #     # Set alignment
    #     for col in sheet.columns:
    #         for cell in col:
    #             cell.alignment = Alignment(horizontal="center", vertical="center")

    #     # Set optimal width
    #     column_widths: list[int] = []
    #     for row in sheet:
    #         for row, cell in enumerate(row):
    #             if len(column_widths) > row:
    #                 if len(str(cell.value)) + 1 > column_widths[row]:
    #                     column_widths[row] = len(str(cell.value)) + 1
    #             else:
    #                 column_widths += [0]
    #     for row, column_width in enumerate(column_widths, start=1):
    #         sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_summary_score_opti(self, workbook: Workbook):
        sheet = workbook.create_sheet("summary")
        # first row
        # first columns are the instances informations then the methods names
        line: list[int | str | float] = [f"/{len(self.instances)}"]
        line += [m for m in self.methods_names]
        sheet.append(line)

        # body of the table

        # nb flat
        line = ["nb flat done"]
        for m in self.methods_names:
            line += [self.nb_flat[m]]
        sheet.append(line)

        # nb solve
        line = ["nb solved"]
        for m in self.methods_names:
            line += [self.nb_solved[m]]
        sheet.append(line)

        # nb best known score
        line = ["nb best known score"]
        for m in self.methods_names:
            line += [self.nb_best_score[m]]
        sheet.append(line)

        # nb optimal
        line = ["nb proved optimal"]
        for m in self.methods_names:
            line += [self.nb_optim[m]]
        sheet.append(line)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) + 1 > column_widths[i]:
                        column_widths[i] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = column_width

    def to_xlsx(self, file_name: str):
        """Convert the table to xlsx file"""
        workbook = Workbook()

        # first sheet with all results of each methods
        print("generation results")
        self.table_results(workbook)

        self.table_results_to_latex(file_name[:-4] + "tex")

        # print("generation gaps")
        # self.table_gaps(workbook)

        print("generation comparison score")
        self.table_comparaison_score(workbook)

        print("generation comparison time")
        self.table_comparaison_time(workbook)

        print("generation comparison optim")
        self.table_comparaison_optim(workbook)

        print("generation comparison optim time")
        self.table_comparaison_optim_time(workbook)

        # print("generation ranks")
        # self.table_ranks(workbook)

        # print("generation summary ranks")
        # self.table_summary_ranks(workbook)

        self.table_summary_score_opti(workbook)

        workbook.save(file_name)


def get_best_known_score(instance: str, problem: str) -> tuple[int, bool]:
    """return best know score in the literature and if score optimal"""
    file: str = f"instances/best_scores_{problem}.txt"
    with open(file, "r", encoding="utf8") as f:
        for line in f.readlines():
            instance_, score, optimal = line[:-1].split(" ")
            if instance_ == instance:
                return int(score), optimal == "*"
    raise Exception(f"instance {instance} not found in {file}")


def get_density(
    instance: str, instance_type: str, problem: str
) -> tuple[int, int, float]:
    """return nb vertices, nb edges and density"""
    info_file = f"instances/instance_info_{problem}.csv"
    with open(info_file, "r", encoding="utf8") as file:
        for line in file.readlines():
            instance_, reduced, nb_vertices, nb_edges, density = line[:-1].split(",")
            is_reduced = reduced == "true"
            if instance_ != instance or (is_reduced and instance_type == "original"):
                continue
            return int(nb_vertices), int(nb_edges), round(float(density), 2)
    print(f"instance {instance} not found in {info_file}")
    return -1, -1, -1


if __name__ == "__main__":
    main()
