"""
Generate a xlsx file containing results of given methods from json files.

Follow the main function to manage the files and parameters
"""
import re
import statistics
from glob import glob
from typing import Any, Iterable
import json
import os

import pandas as pd  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Alignment, Font  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore


# When a best score is equal to the best known score : red
COLOR_BEST = "ff0000"
# When a best score is proven optimal : blue
COLOR_OPTIMAL = "0000ff"
# When a best score is a new proven optimal : blue
COLOR_NEW_OPTIMAL = "00ff00"
# When a best score is better than the best known score : green
COLOR_NEW_BEST = "FF7F00"


COLOR_GAP1 = "488f31"
COLOR_GAP2 = "de425b"


def main():
    """
    Choose the methods and instances and create the xlsx file
    """
    # Add method name and repertory of data of each method
    methods: list[tuple(str, str)] = [
        # E1
        ("dual ortools", "cp_1h_E1_all/E1_dual_ortools"),
        ("primal static h1", "cp_1h_E1_all/E1_primal_static_h1"),
        ("primal static h2", "cp_1h_E1_all/E1_primal_static_h2"),
        ("primal dynamic h1", "cp_1h_E1_all/E1_primal_dynamic_h1"),
        ("primal dynamic h2", "cp_1h_E1_all/E1_primal_dynamic_h2"),
        ("joint static h1", "cp_1h_E1_all/E1_joint_static_h1"),
        ("joint static h2", "cp_1h_E1_all/E1_joint_static_h2"),
        ("joint dyn h1", "cp_1h_E1_all/E1_joint_dynamic_h1"),
        ("joint dyn h2", "cp_1h_E1_all/E1_joint_dynamic_h2"),
    ]

    problem = "wvcp"

    # Choose the set of instances
    instances_set = ("pxx", "pxx")
    instances_set = ("rxx", "rxx")
    instances_set = ("DIMACS_non_optimal", "dimacs_no")
    instances_set = ("DIMACS_optimal", "dimacs_o")
    instances_set = ("../instances_coeff", "instances_coeff")
    instances_set = ("../instances_hard_wvcp", "hard_wvcp")
    instances_set = ("../instances_non_optimal", "non_optimal")
    instances_set = ("../instance_feasible", "feasible")
    instances_set = ("instance_list_wvcp", "all")

    instance_type = "original"
    instance_type = "reduced"

    output_file = f"xlsx_files/E1_1h_{instance_type}_{instances_set[1]}.xlsx"

    with open(f"instances/{instances_set[0]}.txt", "r", encoding="utf8") as file:
        instances = [i[:-1] for i in file.readlines()]

    table = Table(
        methods=methods,
        instances=instances,
        problem=problem,
        instance_type=instance_type,
    )
    table.to_xlsx(output_file)
    print(output_file)


class Method:
    def __init__(
        self, name: str, repertory: str, instance_name, instance_type: str
    ) -> None:
        self.name: str = name
        self.repertory: str = repertory
        self.flat_time: float = float("inf")
        self.solve_time: float = float("inf")
        self.score: float = float("inf")
        self.optimality_time: float = float("inf")
        self.optimal: bool = False
        self.objective_bound: float = float("inf")
        self.failures: float = float("inf")

        # load data
        json_file = f"outputs/{repertory}/{instance_type}_{instance_name}.json"
        cplex_file = f"outputs/{repertory}/{instance_type}_{instance_name}.cplex"
        if os.path.exists(json_file):
            with open(json_file, "r", encoding="utf8") as file:
                for line in file.readlines():
                    l_json = json.loads(line)
                    if l_json["type"] == "statistics":
                        if "flatTime" in l_json["statistics"]:
                            self.flat_time = round(l_json["statistics"]["flatTime"], 1)
                            if self.flat_time > 3600:
                                print(instance_name, "flat time > 3600")
                        else:
                            if self.optimal:
                                self.optimality_time = round(
                                    l_json["statistics"]["solveTime"], 1
                                )
                            else:
                                self.solve_time = round(
                                    l_json["statistics"]["solveTime"], 1
                                )
                            try:
                                self.objective_bound = l_json["statistics"][
                                    "objectiveBound"
                                ]
                            except:
                                pass
                            try:
                                self.failures = l_json["statistics"]["failures"]
                            except:
                                self.objective_bound = round(self.objective_bound, 2)
                    elif l_json["type"] == "solution":
                        if "x_score" in l_json["output"]["json"]:
                            self.score = l_json["output"]["json"]["x_score"]
                        else:
                            self.score = l_json["output"]["json"]["yx_score"]
                    elif l_json["type"] == "status":
                        self.optimal = l_json["status"] == "OPTIMAL_SOLUTION"
        elif os.path.exists(cplex_file):
            self.flat_time = 0
            with open(cplex_file, "r", encoding="utf8") as file:
                for line in file.readlines():
                    split = line.split()
                    if line.startswith("Result"):
                        self.score = int(split[-1])
                    if line.startswith("Total (root+branch&cut)"):
                        self.solve_time = float(split[3])
                    if split and split[-1] == "0.00%":
                        self.optimal = True
                    if line.startswith("All rows and columns eliminated."):
                        self.optimal = True


class Gap:
    def __init__(self, m_1: Method, m_2: Method) -> None:
        inf = float("inf")
        self.score_diff: float = inf
        self.time_diff: float = inf
        self.time_optim_diff = inf
        self.better_method_score: str = ""
        self.better_method_time: str = ""
        self.better_method_optim: str = ""

        if m_1.score == inf and m_2.score == inf:
            return
        if m_1.score == inf:
            self.better_method_score = m_2.name
            return
        if m_2.score == inf:
            self.better_method_score = m_1.name
            return

        self.score_diff = round(m_1.score - m_2.score, 1)
        if self.score_diff < 0:
            self.better_method_score = m_1.name
            return
        if self.score_diff > 0:
            self.better_method_score = m_2.name
            return

        # same score - diff time
        self.time_diff = round(m_1.solve_time - m_2.solve_time, 1)
        if self.time_diff < -1:
            self.better_method_time = m_1.name
        if self.time_diff > 1:
            self.better_method_time = m_2.name

        if m_1.optimality_time == inf and m_2.optimality_time == inf:
            return
        if m_1.optimality_time == inf:
            self.better_method_optim = m_2.name
            return
        if m_2.optimality_time == inf:
            self.better_method_optim = m_1.name
            return

        self.time_optim_diff = round(m_1.optimality_time - m_2.optimality_time, 1)
        if self.time_optim_diff < -1:
            self.better_method_optim = m_1.name
        if self.time_optim_diff > 1:
            self.better_method_optim = m_2.name


class Instance:
    """Store the results of all given method on one instance"""

    def __init__(
        self,
        name: str,
        methods: list[tuple[str, str]],
        problem: str,
        instance_type: str,
        gaps: list[tuple[str, str]],
    ) -> None:
        self.name: str = name
        self.nb_vertices: int
        self.nb_edges: int
        self.best_known_score: int
        self.optimal: bool

        self.methods: dict[str, Method] = {
            m_name: Method(m_name, repertory, name, instance_type)
            for m_name, repertory in methods
        }
        self.methods_names: list[str] = [m_name for m_name, _ in methods]
        self.gaps: dict[tuple[str, str], Gap] = dict()

        # get informations on the instance
        self.nb_vertices, self.nb_edges, self.density = get_density(name, instance_type)
        self.best_known_score, self.optimal = get_best_known_score(name, problem)

        # compute gap between methods
        for m_1, m_2 in gaps:
            self.gaps[(m_1, m_2)] = Gap(self.methods[m_1], self.methods[m_2])


class Table:
    """Representation of the data table"""

    def __init__(
        self,
        methods: list[tuple[str, str]],
        instances: list[str],
        problem: str,
        instance_type: str,
    ) -> None:

        self.methods_names: list[str] = [m_name for m_name, _ in methods]

        self.gaps: list[tuple[str, str]] = []
        # pylint: disable=consider-using-enumerate
        for i in range(len(self.methods_names)):
            for j in range(i + 1, len(self.methods_names)):
                self.gaps.append((self.methods_names[i], self.methods_names[j]))

        self.instances: list[Instance] = [
            Instance(instance, methods, problem, instance_type, self.gaps)
            for instance in instances
        ]

        self.nb_optim = {
            m: sum(1 for instance in self.instances if instance.methods[m].optimal)
            for m in self.methods_names
        }

        self.nb_best_score = {
            m: sum(
                1
                for instance in self.instances
                if instance.methods[m].score == instance.best_known_score
            )
            for m in self.methods_names
        }

        self.nb_gaps_m1_best_score: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_score == m1)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m2_best_score: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_score == m2)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m1_best_time: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_time == m1)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m2_best_time: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_time == m2)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m1_best_optim: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_optim == m1)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m2_best_optim: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_optim == m2)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }

    def __repr__(self) -> str:
        return "\n".join([str(instance) for instance in self.instances])

    def table_results(self, workbook: Workbook):
        sheet = workbook.active
        sheet.title = "results"
        # first row
        # first columns are the instances informations then the methods names
        instance_info = ["instance", "|V|", "|E|", "density", "BKS", "optim"]
        columns_info = [
            "score",
            "flat(s)",
            "solve(s)",
            "opti(s)",
            "LB",
            "failures",
        ]
        line: list[int | str | float] = list(instance_info)
        line += [m for m in self.methods_names for _ in columns_info]
        sheet.append(line)
        # merge first row for methods names
        for i in range(len(self.methods_names)):
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # second row
        # instance informations then columns info
        line = list(instance_info)
        for _ in self.methods_names:
            line += list(columns_info)
        sheet.append(line)

        # merge 2 firsts lines for instances informations
        for i in range(len(instance_info)):
            sheet.merge_cells(
                start_row=1, end_row=2, start_column=i + 1, end_column=i + 1
            )

        # body of the table
        # first columns are the instance info then the scores, times,... for each methods
        for instance in self.instances:
            line = [
                instance.name,
                instance.nb_vertices,
                instance.nb_edges,
                instance.density,
                instance.best_known_score,
                instance.optimal,
            ]
            for m in self.methods_names:
                method = instance.methods[m]
                line += [
                    method.score,
                    method.flat_time,
                    method.solve_time,
                    method.optimality_time,
                    method.objective_bound,
                    method.failures,
                ]
            sheet.append(line)
            for col, m in enumerate(self.methods_names):
                column_best_score = len(instance_info) + 1 + len(columns_info) * col
                cell_best_score = sheet.cell(sheet.max_row, column_best_score)
                if cell_best_score.value == float("inf"):
                    continue
                val_best_score = int(cell_best_score.value)
                if val_best_score == instance.best_known_score:
                    cell_best_score.font = Font(bold=True, color=COLOR_BEST)
                elif val_best_score < instance.best_known_score:
                    cell_best_score.font = Font(bold=True, color=COLOR_NEW_BEST)
                if instance.methods[m].optimal and instance.optimal:
                    cell_best_score.font = Font(bold=True, color=COLOR_OPTIMAL)
                elif instance.methods[m].optimal:
                    cell_best_score.font = Font(bold=True, color=COLOR_NEW_OPTIMAL)

        # footer

        # nb optimal
        line = ["nb proved optimal"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_optim[m]}/{len(self.instances)}"] * len(columns_info)
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # nb best known score
        line = ["nb best known score"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_best_score[m]}/{len(self.instances)}"] * len(
                columns_info
            )
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) + 1 > column_widths[i]:
                        column_widths[i] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = column_width

        # Freeze row and columns
        sheet.freeze_panes = sheet["G3"]

    def table_gaps(self, workbook: Workbook):
        sheet = workbook.create_sheet("gaps")

        instance_info = ["instance"]
        columns_info = ["gap score", "gap time", "gap optim"]
        # first row
        line: list[int | str | float] = list(instance_info)
        line += [
            f"m1 : {m1} \n- m2 : {m2}" for m1, m2 in self.gaps for _ in columns_info
        ]
        sheet.append(line)
        # merge first row for methods names
        for i in range(len(self.gaps)):
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # second row
        # instance informations then columns info
        line = list(instance_info)
        for i in range(len(self.gaps)):
            line += list(columns_info)
        sheet.append(line)

        # merge 2 firsts lines for instances informations
        for i in range(len(instance_info)):
            sheet.merge_cells(
                start_row=1, end_row=2, start_column=i + 1, end_column=i + 1
            )

        # body of the table
        # first columns are the instance info then the gaps for each methods
        for instance in self.instances:
            line = [instance.name]
            for m1, m2 in self.gaps:
                gap = instance.gaps[(m1, m2)]
                line += [
                    gap.score_diff,
                    gap.time_diff,
                    gap.time_optim_diff,
                ]
            sheet.append(line)

            # add color
            for col in range(len(self.gaps)):
                column_diff_score = len(instance_info) + 1 + len(columns_info) * col
                column_diff_time = column_diff_score + 1
                column_diff_optim = column_diff_time + 1
                cell_diff_score = sheet.cell(sheet.max_row, column_diff_score)
                cell_diff_time = sheet.cell(sheet.max_row, column_diff_time)
                cell_diff_optim = sheet.cell(sheet.max_row, column_diff_optim)
                if cell_diff_score.value == float("inf"):
                    continue
                val_diff_score = float(cell_diff_score.value)
                val_diff_time = float(cell_diff_time.value)
                val_diff_optim = float(cell_diff_optim.value)

                color_score = "808080"  # gray
                color_time = "808080"  # gray
                color_optim = "808080"  # gray
                if val_diff_score > 0:
                    # 1 better
                    color_score = COLOR_GAP1
                elif val_diff_score < 0:
                    # 2 better
                    color_score = COLOR_GAP2
                else:
                    if val_diff_time > 1:
                        color_time = COLOR_GAP1
                    elif val_diff_time < -1:
                        color_time = COLOR_GAP2
                    if val_diff_optim > 1:
                        color_optim = COLOR_GAP1
                    elif val_diff_optim < -1:
                        color_optim = COLOR_GAP2
                cell_diff_score.font = Font(bold=True, color=color_score)
                cell_diff_time.font = Font(bold=True, color=color_time)
                cell_diff_optim.font = Font(bold=True, color=color_optim)

        # footer
        line = ["m1 better"] * len(instance_info)
        for m1, m2 in self.gaps:
            line += [
                self.nb_gaps_m1_best_score[(m1, m2)],
                self.nb_gaps_m1_best_time[(m1, m2)],
                self.nb_gaps_m1_best_optim[(m1, m2)],
            ]
        sheet.append(line)
        for col in range(1, sheet.max_column + 1):
            sheet.cell(sheet.max_row, col).font = Font(bold=True, color=COLOR_GAP2)

        line = ["m2 better"] * len(instance_info)
        for m1, m2 in self.gaps:
            line += [
                self.nb_gaps_m2_best_score[(m1, m2)],
                self.nb_gaps_m2_best_time[(m1, m2)],
                self.nb_gaps_m2_best_optim[(m1, m2)],
            ]
        sheet.append(line)

        for col in range(1, sheet.max_column + 1):
            sheet.cell(sheet.max_row, col).font = Font(bold=True, color=COLOR_GAP1)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) + 1 > column_widths[i]:
                        column_widths[i] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = column_width

        # Freeze row and columns
        sheet.freeze_panes = sheet["B3"]

    def table_comparaison_score(self, workbook: Workbook):
        sheet = workbook.create_sheet("score")
        # first row with the methods names
        line: list[str | int] = [""]
        for m1 in self.methods_names:
            line.append(m1)
        sheet.append(line)

        # add data, first method name then nb significative gaps
        for row, m1 in enumerate(self.methods_names):
            m1 = self.methods_names[row]
            line = []
            line.append(m1)
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    line.append("-")
                else:
                    if row < col:
                        line.append(self.nb_gaps_m1_best_score[(m1, m2)])
                    else:
                        line.append(self.nb_gaps_m2_best_score[(m2, m1)])
            sheet.append(line)

            # bold best scores
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    continue
                cell_m1 = sheet.cell(row + 2, col + 2)
                if (
                    row < col
                    and (
                        self.nb_gaps_m1_best_score[(m1, m2)]
                        > self.nb_gaps_m2_best_score[(m1, m2)]
                    )
                ) or (
                    row > col
                    and (
                        self.nb_gaps_m2_best_score[(m2, m1)]
                        > self.nb_gaps_m1_best_score[(m2, m1)]
                    )
                ):
                    cell_m1.font = Font(bold=True)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rotate the first row
        for col in range(len(self.methods_names)):
            met_name_cell = sheet.cell(1, col + 2)
            met_name_cell.alignment = Alignment(textRotation=90)

        # add info
        line = [""] * (len(self.methods_names) + 1)
        line.append("number of times the method on the row is find a better solution")
        sheet.append(line)

        # Set optimal width
        column_widths: list[int] = []
        row_: Iterable[Iterable[Any]]
        for row_ in sheet:
            for col, cell in enumerate(row_):
                if len(column_widths) > col:
                    if len(str(cell.value)) + 1 > column_widths[col]:
                        column_widths[col] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_comparaison_time(self, workbook: Workbook):
        sheet = workbook.create_sheet("time")
        # first row with the methods names
        line: list[str | int] = [""]
        for m1 in self.methods_names:
            line.append(m1)
        sheet.append(line)

        # add data, first method name then nb significative gaps
        for row, m1 in enumerate(self.methods_names):
            m1 = self.methods_names[row]
            line = []
            line.append(m1)
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    line.append("-")
                else:
                    if row < col:
                        line.append(self.nb_gaps_m1_best_time[(m1, m2)])
                    else:
                        line.append(self.nb_gaps_m2_best_time[(m2, m1)])
            sheet.append(line)

            # bold best scores
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    continue
                cell_m1 = sheet.cell(row + 2, col + 2)
                if (
                    row < col
                    and (
                        self.nb_gaps_m1_best_time[(m1, m2)]
                        > self.nb_gaps_m2_best_time[(m1, m2)]
                    )
                ) or (
                    row > col
                    and (
                        self.nb_gaps_m2_best_time[(m2, m1)]
                        > self.nb_gaps_m1_best_time[(m2, m1)]
                    )
                ):
                    cell_m1.font = Font(bold=True)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rotate the first row
        for col in range(len(self.methods_names)):
            met_name_cell = sheet.cell(1, col + 2)
            met_name_cell.alignment = Alignment(textRotation=90)

        # add info
        line = [""] * (len(self.methods_names) + 1)
        line.append(
            "number of times the method on the row is faster to find best solution (if they found the same score)"
        )
        sheet.append(line)

        # Set optimal width
        column_widths: list[int] = []
        row_: Iterable[Iterable[Any]]
        for row_ in sheet:
            for col, cell in enumerate(row_):
                if len(column_widths) > col:
                    if len(str(cell.value)) + 1 > column_widths[col]:
                        column_widths[col] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_comparaison_optim(self, workbook: Workbook):
        sheet = workbook.create_sheet("optim")
        # first row with the methods names
        line: list[str | int] = [""]
        for m1 in self.methods_names:
            line.append(m1)
        sheet.append(line)

        # add data, first method name then nb significative gaps
        for row, m1 in enumerate(self.methods_names):
            m1 = self.methods_names[row]
            line = []
            line.append(m1)
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    line.append("-")
                else:
                    if row < col:
                        line.append(self.nb_gaps_m1_best_optim[(m1, m2)])
                    else:
                        line.append(self.nb_gaps_m2_best_optim[(m2, m1)])
            sheet.append(line)

            # bold best scores
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    continue
                cell_m1 = sheet.cell(row + 2, col + 2)
                if (
                    row < col
                    and (
                        self.nb_gaps_m1_best_optim[(m1, m2)]
                        > self.nb_gaps_m2_best_optim[(m1, m2)]
                    )
                ) or (
                    row > col
                    and (
                        self.nb_gaps_m2_best_optim[(m2, m1)]
                        > self.nb_gaps_m1_best_optim[(m2, m1)]
                    )
                ):
                    cell_m1.font = Font(bold=True)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rotate the first row
        for col in range(len(self.methods_names)):
            met_name_cell = sheet.cell(1, col + 2)
            met_name_cell.alignment = Alignment(textRotation=90)

        # add info
        line = [""] * (len(self.methods_names) + 1)
        line.append(
            "number of times the method on the row is faster to prove optimality (if they found the same score)"
        )
        sheet.append(line)

        # Set optimal width
        column_widths: list[int] = []
        row_: Iterable[Iterable[Any]]
        for row_ in sheet:
            for col, cell in enumerate(row_):
                if len(column_widths) > col:
                    if len(str(cell.value)) + 1 > column_widths[col]:
                        column_widths[col] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def to_xlsx(self, file_name: str):
        """Convert the table to xlsx file"""
        workbook = Workbook()

        # first sheet with all results of each methods
        print("generation results")
        self.table_results(workbook)

        print("generation gaps")
        self.table_gaps(workbook)

        print("generation comparison score")
        self.table_comparaison_score(workbook)

        print("generation comparison time")
        self.table_comparaison_time(workbook)

        print("generation comparison optim")
        self.table_comparaison_optim(workbook)

        workbook.save(file_name)


def get_best_known_score(instance: str, problem: str) -> tuple[int, bool]:
    """return best know score in the literature and if score optimal"""
    file: str = f"instances/best_scores_{problem}.txt"
    with open(file, "r", encoding="utf8") as f:
        for line in f.readlines():
            instance_, score, optimal = line[:-1].split(" ")
            if instance_ == instance:
                return int(score), optimal == "*"
    raise Exception(f"instance {instance} not found in {file}")


def get_density(instance: str, instance_type: str) -> tuple[int, int, float]:
    """return nb vertices, nb edges and density"""
    info_file = f"density_{instance_type}.csv"
    with open(info_file, "r", encoding="utf8") as file:
        for line in file.readlines():
            instance_, nb_vertices, nb_edges, density = line[:-1].split(",")
            if instance_ != instance:
                continue
            return int(nb_vertices), int(nb_edges), round(float(density), 2)
    print(f"instance {instance} not found in {info_file}")
    return -1, -1, -1


if __name__ == "__main__":
    main()
